# Re-run creation after kernel reset (rebuild both files).
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def create_detailed_file(path):
    wb = Workbook()
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    ws_inputs.append(["Parameter (EN)", "参数中文说明", "Value", "Notes"])
    inputs = [
        ("roof_total_area_m2", "屋顶总面积 (m²)", 120.0, "来自上游 3D 还原"),
        ("roof_effective_area_m2", "屋顶有效面积 (m²)", 80.0, "过滤朝向/过小坡面后的面积和"),
        ("roof_max_panels", "屋顶最大面板数 (块)", 16, "上游计算得出"),
        ("panel_area_m2", "单块面板面积 (m²)", 1.94, "面板规格"),
        ("panel_power_kw", "单块面板功率 (kW)", 0.41, "面板标称功率"),
        ("panel_unit_cost", "单块面板成本 (AUD/块)", 350, ""),
        ("inverter_unit_cost_per_kw", "逆变器成本 (AUD/kW)", 250, ""),
        ("battery_unit_cost_per_kwh", "电池成本 (AUD/kWh)", 800, ""),
        ("install_base_cost", "安装基础费用 (AUD)", 1000, ""),
        ("install_cost_per_kw", "安装每kW费用 (AUD/kW)", 1000, ""),
        ("dc_ac_ratio", "容配比 (DC/AC)", 1.33, ""),
        ("yield_per_kw_per_year", "每kW年发电量 (kWh/kW/年)", 1460, "兜底值，pvlib 优先"),
        ("baseline_self_consumption_rate", "基线自用率 (无电池)", 0.30, ""),
        ("plan_a_capacity_factor", "方案A 容量系数", 0.2, ""),
        ("plan_b_capacity_factor", "方案B 容量系数", 0.5, ""),
        ("plan_c_capacity_factor", "方案C 容量系数", 0.8, ""),
        ("plan_d_capacity_factor", "方案D 容量系数", 1.0, ""),
        ("plan_a_min_kw", "方案A 最小装机下限 (kW)", 3.5, ""),
        ("plan_c_max_kw", "方案C 最大上限 (kW)", 13.3, ""),
        ("plan_a_target_sc_rate", "方案A 目标自用率", 0.35, ""),
        ("plan_b_target_sc_rate", "方案B 目标自用率", 0.45, ""),
        ("plan_c_target_sc_rate", "方案C 目标自用率", 0.50, ""),
        ("plan_d_target_sc_rate", "方案D 目标自用率", 0.60, ""),
        ("battery_dod", "电池放电深度 DoD", 0.90, ""),
        ("battery_rte", "电池往返效率 RTE", 0.90, ""),
        ("battery_install_base_cost", "电池安装基础费 (AUD)", 500, ""),
        ("battery_install_cost_per_kwh", "电池安装每kWh费用 (AUD/kWh)", 50, ""),
        ("battery_effective_usage_factor", "电池有效使用系数", 0.90, ""),
        ("profit_margin_rate", "利润率", 0.10, ""),
        ("price_range_percent", "报价浮动范围", 0.05, ""),
        ("grid_buy_rate", "购电价 (AUD/kWh)", 0.33, ""),
        ("grid_sell_rate", "售电/馈网价 (AUD/kWh)", 0.07, ""),
        ("annual_home_usage_proxy_low", "Low 用户年用电 (kWh)", 3000, ""),
        ("annual_home_usage_proxy_med", "Med 用户年用电 (kWh)", 6000, ""),
        ("annual_home_usage_proxy_high", "High 用户年用电 (kWh)", 9000, ""),
        ("existing_solar_annual_gen_kwh", "已有光伏估算年发电量 (kWh/年)", "", "留空则使用估算"),
        ("existing_sc_rate", "已有系统基线自用率", 0.30, ""),
        ("retrofit_plan_a_kwh", "Retrofit A 电池容量 (kWh)", 5, ""),
        ("retrofit_plan_b_kwh", "Retrofit B 电池容量 (kWh)", 10, ""),
        ("retrofit_plan_c_kwh", "Retrofit C 电池容量 (kWh)", 13.5, ""),
        ("retrofit_plan_d_kwh", "Retrofit D 电池容量 (kWh)", 20, ""),
    ]
    row = 2
    refs = {}
    for key, cn, val, note in inputs:
        ws_inputs.append([key, cn, val, note])
        refs[key] = f"C{row}"
        row += 1
    for cell in ws_inputs["1:1"]:
        cell.font = Font(bold=True)
    ws_inputs.append([])
    ws_inputs.append(["说明：roof_max_panels 来自上游面积计算，panel_power_kw 与 panel_area_m2 为面板规格。pvlib 优先用于发电量计算，若不可用则使用 yield_per_kw_per_year。"])

    ws_plans = wb.create_sheet("Plans_By_Columns")
    ws_plans.append(["参数（中/EN）", "Plan A", "Plan B", "Plan C", "Plan D"])
    for cell in ws_plans["1:1"]:
        cell.font = Font(bold=True)
    plan_params = [
        ("solar_kw", "光伏装机容量 (kW)"),
        ("panel_count", "面板数量 (块)"),
        ("inverter_kw", "逆变器容量 (kW)"),
        ("annual_generation_kwh", "年发电量 (kWh/年)"),
        ("daily_energy_to_shift_kwh", "每日需转移能量 (kWh/天)"),
        ("battery_nominal_kwh", "建议电池标称容量 (kWh)"),
        ("battery_pack_suggested_kwh", "推荐商用电池规格 (kWh)"),
        ("cost_panels", "面板成本 (AUD)"),
        ("cost_inverter", "逆变器成本 (AUD)"),
        ("cost_battery", "电池成本 (AUD)"),
        ("cost_install", "安装成本 (AUD)"),
        ("total_cost", "项目总成本 (AUD)"),
        ("price_base", "报价基准 (AUD)"),
        ("payback_base_years", "回本年限 基线 (年)"),
        ("payback_low_years", "回本年限 保守 (年)"),
        ("payback_high_years", "回本年限 乐观 (年)"),
    ]
    r = 2
    for key, cn in plan_params:
        ws_plans[f"A{r}"] = f"{cn}\n({key})"
        r += 1
    param_row = {plan_params[i][0]: 2+i for i in range(len(plan_params))}
    srefs = {k: "Inputs!" + v for k, v in refs.items()}

    def set_plan(col_idx, plan_letter):
        col = get_column_letter(col_idx)
        row_solar = param_row["solar_kw"]
        if plan_letter == "A":
            capf = srefs['plan_a_capacity_factor']
            min_kw = srefs['plan_a_min_kw']
            formula = f"=MAX(MIN({srefs['roof_max_panels']}*{srefs['panel_power_kw']}*{capf}, {srefs['plan_c_max_kw']}), {min_kw})"
        elif plan_letter == "B":
            capf = srefs['plan_b_capacity_factor']
            formula = f"=MIN({srefs['roof_max_panels']}*{srefs['panel_power_kw']}*{capf}, {srefs['plan_c_max_kw']})"
        elif plan_letter == "C":
            capf = srefs['plan_c_capacity_factor']
            formula = f"=MIN({srefs['roof_max_panels']}*{srefs['panel_power_kw']}*{capf}, {srefs['plan_c_max_kw']})"
        else:
            capf = srefs['plan_d_capacity_factor']
            formula = f"=MIN({srefs['roof_max_panels']}*{srefs['panel_power_kw']}*{capf}, {srefs['plan_c_max_kw']})"
        ws_plans[f"{col}{row_solar}"] = formula
        # panel_count floor
        r_panel = param_row["panel_count"]
        ws_plans[f"{col}{r_panel}"] = f"=INT({col}{row_solar}/{srefs['panel_power_kw']})"
        # inverter_kw = CEILING(solar_kw / dc_ac_ratio, 0.1)
        r_inv = param_row["inverter_kw"]
        ws_plans[f"{col}{r_inv}"] = f"=CEILING({col}{row_solar}/{srefs['dc_ac_ratio']},0.1)"
        # annual_generation_kwh fallback
        r_gen = param_row["annual_generation_kwh"]
        ws_plans[f"{col}{r_gen}"] = f"={col}{row_solar}*{srefs['yield_per_kw_per_year']}"
        # daily_energy_to_shift
        r_daily = param_row["daily_energy_to_shift_kwh"]
        if plan_letter == "A":
            target = srefs['plan_a_target_sc_rate']
        elif plan_letter == "B":
            target = srefs['plan_b_target_sc_rate']
        elif plan_letter == "C":
            target = srefs['plan_c_target_sc_rate']
        else:
            target = srefs['plan_d_target_sc_rate']
        ws_plans[f"{col}{r_daily}"] = f"=({col}{r_gen}/365)*({target}-{srefs['baseline_self_consumption_rate']})"
        # battery_nominal_kwh = CEILING(...,1)
        r_batt = param_row["battery_nominal_kwh"]
        ws_plans[f"{col}{r_batt}"] = f"=IF({col}{r_daily}>0, CEILING({col}{r_daily}/({srefs['battery_dod']}*{srefs['battery_rte']}),1), 0)"
        # suggested commercial pack
        r_pack = param_row["battery_pack_suggested_kwh"]
        ws_plans[f"{col}{r_pack}"] = f"=IF({col}{r_batt}>={20},20,IF({col}{r_batt}>={13.5},13.5,IF({col}{r_batt}>={10},10,IF({col}{r_batt}>={6.5},6.5,5))))"
        # costs
        r_cp = param_row["cost_panels"]
        ws_plans[f"{col}{r_cp}"] = f"={col}{r_panel}*{srefs['panel_unit_cost']}"
        r_ci = param_row["cost_inverter"]
        ws_plans[f"{col}{r_ci}"] = f"={col}{r_inv}*{srefs['inverter_unit_cost_per_kw']}"
        r_cb = param_row["cost_battery"]
        ws_plans[f"{col}{r_cb}"] = f"={col}{r_batt}*{srefs['battery_unit_cost_per_kwh']}"
        r_cins = param_row["cost_install"]
        ws_plans[f"{col}{r_cins}"] = f"={srefs['install_base_cost']} + ({col}{row_solar}*{srefs['install_cost_per_kw']})"
        r_tot = param_row["total_cost"]
        ws_plans[f"{col}{r_tot}"] = f"={col}{r_cp}+{col}{r_ci}+{col}{r_cb}+{col}{r_cins}"
        r_pb = param_row["price_base"]
        ws_plans[f"{col}{r_pb}"] = f"={col}{r_tot}*(1+{srefs['profit_margin_rate']})"
        r_pay_base = param_row["payback_base_years"]
        r_pay_low = param_row["payback_low_years"]
        r_pay_high = param_row["payback_high_years"]
        ws_plans[f"{col}{r_pay_base}"] = f"=IFERROR(({col}{r_pb}) / ( (MIN({col}{r_gen}*{target},{srefs['annual_home_usage_proxy_med']})*{srefs['grid_buy_rate']}) + (({col}{r_gen} - MIN({col}{r_gen}*{target},{srefs['annual_home_usage_proxy_med']}))*{srefs['grid_sell_rate']}) ), \"Inf\")"
        ws_plans[f"{col}{r_pay_low}"] = f"=IFERROR( ({col}{r_tot}*1.1*(1+{srefs['profit_margin_rate']})) / ( (MIN({col}{r_gen}*0.85*{target},{srefs['annual_home_usage_proxy_med']})*{srefs['grid_buy_rate']}*0.8) + (({col}{r_gen}*0.85 - MIN({col}{r_gen}*0.85*{target},{srefs['annual_home_usage_proxy_med']}))*{srefs['grid_sell_rate']}) ), \"Inf\")"
        ws_plans[f"{col}{r_pay_high}"] = f"=IFERROR( ({col}{r_tot}*0.9*(1+{srefs['profit_margin_rate']})) / ( (MIN({col}{r_gen}*1.15*{target},{srefs['annual_home_usage_proxy_med']})*{srefs['grid_buy_rate']}*1.2) + (({col}{r_gen}*1.15 - MIN({col}{r_gen}*1.15*{target},{srefs['annual_home_usage_proxy_med']}))*{srefs['grid_sell_rate']}) ), \"Inf\")"

    for idx, pl in enumerate(["A","B","C","D"], start=2):
        set_plan(idx, pl)

    ws_plans[f"A{2+len(plan_params)}"] = "说明：panel_count 使用向下取整(INT)，逆变器容量向上取整到 0.1 kW，电池标称向上取整到整数 kWh，并建议商用规格。"

    ws_retro = wb.create_sheet("Battery_Retrofit")
    ws_retro.append(["参数（中/EN）", "Retrofit A", "Retrofit B", "Retrofit C", "Retrofit D"])
    for cell in ws_retro["1:1"]:
        cell.font = Font(bold=True)
    retro_params = [
        ("battery_nominal_kwh", "电池标称容量 (kWh)"),
        ("usable_battery_capacity_kwh", "可用电池容量 (kWh)"),
        ("annual_shifted_kwh_est", "全年可转移电量 估计 (kWh/年)"),
        ("max_shiftable_kwh", "可最大转移上限 (kWh/年)"),
        ("final_annual_shifted_kwh", "实际可转移电量 (kWh/年)"),
        ("annual_savings", "年度节省 (AUD/年)"),
        ("total_cost", "项目总成本 (AUD)"),
        ("price_base", "报价基准 (AUD)"),
        ("payback_years", "回本年限 (年)"),
        ("new_self_consumption_rate", "加电池后自用率"),
        ("roi_warning", "回报警告"),
    ]
    r = 2
    for key, cn in retro_params:
        ws_retro[f"A{r}"] = f"{cn}\n({key})"
        r += 1
    sizes = [srefs['retrofit_plan_a_kwh'], srefs['retrofit_plan_b_kwh'], srefs['retrofit_plan_c_kwh'], srefs['retrofit_plan_d_kwh']]
    esg = srefs['existing_solar_annual_gen_kwh']
    for col_idx in range(2,6):
        col = get_column_letter(col_idx)
        size = sizes[col_idx-2]
        ws_retro[f"{col}2"] = f"={size}"
        ws_retro[f"{col}3"] = f"={col}2*{srefs['battery_dod']}"
        ws_retro[f"{col}4"] = f"={col}3*{srefs['battery_effective_usage_factor']}*365"
        ws_retro[f"{col}5"] = f"=IF({esg}=\"\", ({srefs['roof_max_panels']}*{srefs['panel_power_kw']}*0.7*1200)*(1-{srefs['existing_sc_rate']}), {esg}*(1-{srefs['existing_sc_rate']}))"
        ws_retro[f"{col}6"] = f"=MIN({col}4, {col}5)"
        ws_retro[f"{col}7"] = f"={col}6*({srefs['grid_buy_rate']}-{srefs['grid_sell_rate']})"
        ws_retro[f"{col}8"] = f"={col}2*{srefs['battery_unit_cost_per_kwh']} + {srefs['battery_install_base_cost']} + ({col}2*{srefs['battery_install_cost_per_kwh']})"
        ws_retro[f"{col}9"] = f"={col}8*(1+{srefs['profit_margin_rate']})"
        ws_retro[f"{col}10"] = f"=IF({col}7>0, {col}9/{col}7, \"Inf\")"
        ws_retro[f"{col}11"] = f"=IF(OR({esg}=\"\", {esg}=0), (({srefs['roof_max_panels']}*{srefs['panel_power_kw']}*0.7*1200)*{srefs['existing_sc_rate']} + {col}6)/({srefs['roof_max_panels']}*{srefs['panel_power_kw']}*0.7*1200), ({esg}*{srefs['existing_sc_rate']} + {col}6)/{esg})"
        ws_retro[f"{col}12"] = f"=IF({col}6<500, \"Low ROI - small export available\", \"OK\")"

    ws_readme = wb.create_sheet("Readme_Chinese")
    ws_readme.append(["说明（中文）"])
    ws_readme.append(["此版本为详细版，保留面板/逆变器/电池分项成本与取整逻辑。"])
    for cell in ws_readme["1:1"]:
        cell.font = Font(bold=True)

    for ws in [ws_inputs, ws_plans, ws_retro]:
        for col in range(1, ws.max_column+1):
            max_len = 0
            col_letter = get_column_letter(col)
            for cell in ws[col_letter]:
                try:
                    l = len(str(cell.value))
                except:
                    l = 0
                if l > max_len:
                    max_len = l
            ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(path)
    return path

def create_simplified_file(path):
    wb = Workbook()
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    ws_inputs.append(["Parameter (EN)", "参数中文说明", "Value", "Notes"])
    inputs = [
        ("roof_total_area_m2", "屋顶总面积 (m²)", 120.0, "来自上游 3D 还原"),
        ("roof_effective_area_m2", "屋顶有效面积 (m²)", 80.0, "过滤朝向/过小坡面后的面积和"),
        ("roof_max_panels", "屋顶最大面板数 (块)", 16, "上游计算得出"),
        ("panel_area_m2", "单块面板面积 (m²)", 1.94, "面板规格"),
        ("panel_power_kw", "单块面板功率 (kW)", 0.41, "面板标称功率"),
        ("hardware_cost_per_kw", "硬件成本 (AUD/kW)", 1200, "包含面板+逆变器等硬件的简化包价"),
        ("battery_unit_cost_per_kwh", "电池成本 (AUD/kWh)", 800, ""),
        ("install_base_cost", "安装基础费用 (AUD)", 1000, ""),
        ("install_cost_per_kw", "安装每kW费用 (AUD/kW)", 1000, ""),
        ("dc_ac_ratio", "容配比 (DC/AC)", 1.33, ""),
        ("yield_per_kw_per_year", "每kW年发电量 (kWh/kW/年)", 1460, "兜底值，pvlib 优先"),
        ("baseline_self_consumption_rate", "基线自用率 (无电池)", 0.30, ""),
        ("plan_a_capacity_factor", "方案A 容量系数", 0.2, ""),
        ("plan_b_capacity_factor", "方案B 容量系数", 0.5, ""),
        ("plan_c_capacity_factor", "方案C 容量系数", 0.8, ""),
        ("plan_d_capacity_factor", "方案D 容量系数", 1.0, ""),
        ("plan_a_min_kw", "方案A 最小装机下限 (kW)", 3.5, ""),
        ("plan_c_max_kw", "方案C 最大上限 (kW)", 13.3, ""),
        ("plan_a_target_sc_rate", "方案A 目标自用率", 0.35, ""),
        ("plan_b_target_sc_rate", "方案B 目标自用率", 0.45, ""),
        ("plan_c_target_sc_rate", "方案C 目标自用率", 0.50, ""),
        ("plan_d_target_sc_rate", "方案D 目标自用率", 0.60, ""),
        ("battery_dod", "电池放电深度 DoD", 0.90, ""),
        ("battery_rte", "电池往返效率 RTE", 0.90, ""),
        ("battery_install_base_cost", "电池安装基础费 (AUD)", 500, ""),
        ("battery_install_cost_per_kwh", "电池安装每kWh费用 (AUD/kWh)", 50, ""),
        ("battery_effective_usage_factor", "电池有效使用系数", 0.90, ""),
        ("profit_margin_rate", "利润率", 0.10, ""),
        ("price_range_percent", "报价浮动范围", 0.05, ""),
        ("grid_buy_rate", "购电价 (AUD/kWh)", 0.33, ""),
        ("grid_sell_rate", "售电/馈网价 (AUD/kWh)", 0.07, ""),
        ("annual_home_usage_proxy_med", "Med 用户年用电 (kWh)", 6000, ""),
        ("existing_solar_annual_gen_kwh", "已有光伏估算年发电量 (kWh/年)", "", "留空则使用估算"),
        ("existing_sc_rate", "已有系统基线自用率", 0.30, ""),
        ("retrofit_plan_a_kwh", "Retrofit A 电池容量 (kWh)", 5, ""),
        ("retrofit_plan_b_kwh", "Retrofit B 电池容量 (kWh)", 10, ""),
        ("retrofit_plan_c_kwh", "Retrofit C 电池容量 (kWh)", 13.5, ""),
        ("retrofit_plan_d_kwh", "Retrofit D 电池容量 (kWh)", 20, ""),
    ]
    refs = {}
    row = 2
    for key, cn, val, note in inputs:
        ws_inputs.append([key, cn, val, note])
        refs[key] = f"C{row}"
        row += 1
    for cell in ws_inputs["1:1"]:
        cell.font = Font(bold=True)
    ws_inputs.append([])
    ws_inputs.append(["说明：此版本使用 hardware_cost_per_kw 简化硬件成本计算，适合快速报价场景。"])

    ws_plans = wb.create_sheet("Plans_Simplified")
    ws_plans.append(["参数（中/EN）", "Plan A", "Plan B", "Plan C", "Plan D"])
    for cell in ws_plans["1:1"]:
        cell.font = Font(bold=True)
    plan_params = [
        ("solar_kw", "光伏装机容量 (kW)"),
        ("panel_count", "面板数量 (块)"),
        ("annual_generation_kwh", "年发电量 (kWh/年)"),
        ("total_hardware_cost", "硬件成本 (AUD)"),
        ("cost_battery", "电池成本 (AUD)"),
        ("cost_install", "安装成本 (AUD)"),
        ("total_cost", "项目总成本 (AUD)"),
        ("price_base", "报价基准 (AUD)"),
        ("payback_base_years", "回本年限 基线 (年)"),
        ("payback_low_years", "回本年限 保守 (年)"),
        ("payback_high_years", "回本年限 乐观 (年)"),
    ]
    r = 2
    for key, cn in plan_params:
        ws_plans[f"A{r}"] = f"{cn}\n({key})"
        r += 1
    param_row = {plan_params[i][0]: 2+i for i in range(len(plan_params))}
    srefs = {k: "Inputs!" + v for k, v in refs.items()}

    def set_plan_simple(col_idx, plan_letter):
        col = get_column_letter(col_idx)
        row_solar = param_row["solar_kw"]
        if plan_letter == "A":
            formula = f"=MAX(MIN({srefs['roof_max_panels']}*{srefs['panel_power_kw']}*{srefs['plan_a_capacity_factor']}, {srefs['plan_c_max_kw']}), {srefs['plan_a_min_kw']})"
        elif plan_letter == "B":
            formula = f"=MIN({srefs['roof_max_panels']}*{srefs['panel_power_kw']}*{srefs['plan_b_capacity_factor']}, {srefs['plan_c_max_kw']})"
        elif plan_letter == "C":
            formula = f"=MIN({srefs['roof_max_panels']}*{srefs['panel_power_kw']}*{srefs['plan_c_capacity_factor']}, {srefs['plan_c_max_kw']})"
        else:
            formula = f"=MIN({srefs['roof_max_panels']}*{srefs['panel_power_kw']}*{srefs['plan_d_capacity_factor']}, {srefs['plan_c_max_kw']})"
        ws_plans[f"{col}{row_solar}"] = formula
        r_panel = param_row["panel_count"]
        ws_plans[f"{col}{r_panel}"] = f"=INT({col}{row_solar}/{srefs['panel_power_kw']})"
        r_gen = param_row["annual_generation_kwh"]
        ws_plans[f"{col}{r_gen}"] = f"={col}{row_solar}*{srefs['yield_per_kw_per_year']}"
        r_hw = param_row["total_hardware_cost"]
        ws_plans[f"{col}{r_hw}"] = f"={col}{row_solar}*{srefs['hardware_cost_per_kw']}"
        r_cb = param_row["cost_battery"]
        ws_plans[f"{col}{r_cb}"] = "=0"
        r_ci = param_row["cost_install"]
        ws_plans[f"{col}{r_ci}"] = f"={srefs['install_base_cost']} + ({col}{row_solar}*{srefs['install_cost_per_kw']})"
        r_tot = param_row["total_cost"]
        ws_plans[f"{col}{r_tot}"] = f"={col}{r_hw}+{col}{r_cb}+{col}{r_ci}"
        r_pb = param_row["price_base"]
        ws_plans[f"{col}{r_pb}"] = f"={col}{r_tot}*(1+{srefs['profit_margin_rate']})"
        r_pay_base = param_row["payback_base_years"]
        target = srefs['plan_b_target_sc_rate'] if plan_letter=="B" else (srefs['plan_c_target_sc_rate'] if plan_letter=="C" else (srefs['plan_d_target_sc_rate'] if plan_letter=="D" else srefs['plan_a_target_sc_rate']))
        ws_plans[f"{col}{r_pay_base}"] = f"=IFERROR({col}{r_pb}/((MIN({col}{r_gen}*{target},{srefs['annual_home_usage_proxy_med']})*{srefs['grid_buy_rate']}) + (({col}{r_gen} - MIN({col}{r_gen}*{target},{srefs['annual_home_usage_proxy_med']}))*{srefs['grid_sell_rate']})), \"Inf\")"
        ws_plans[f"{col}{param_row['payback_low_years']}"] = f"=IFERROR(({col}{r_tot}*1.1*(1+{srefs['profit_margin_rate']})) / ( (MIN({col}{r_gen}*0.85*{target},{srefs['annual_home_usage_proxy_med']})*{srefs['grid_buy_rate']}*0.8) + (({col}{r_gen}*0.85 - MIN({col}{r_gen}*0.85*{target},{srefs['annual_home_usage_proxy_med']}))*{srefs['grid_sell_rate']}) ), \"Inf\")"
        ws_plans[f"{col}{param_row['payback_high_years']}"] = f"=IFERROR(({col}{r_tot}*0.9*(1+{srefs['profit_margin_rate']})) / ( (MIN({col}{r_gen}*1.15*{target},{srefs['annual_home_usage_proxy_med']})*{srefs['grid_buy_rate']}*1.2) + (({col}{r_gen}*1.15 - MIN({col}{r_gen}*1.15*{target},{srefs['annual_home_usage_proxy_med']}))*{srefs['grid_sell_rate']}) ), \"Inf\")"

    for idx, pl in enumerate(["A","B","C","D"], start=2):
        set_plan_simple(idx, pl)

    ws_retro = wb.create_sheet("Battery_Retrofit")
    ws_retro.append(["参数（中/EN）", "Retrofit A", "Retrofit B", "Retrofit C", "Retrofit D"])
    for cell in ws_retro["1:1"]:
        cell.font = Font(bold=True)
    retro_params = [
        ("battery_nominal_kwh", "电池标称容量 (kWh)"),
        ("usable_battery_capacity_kwh", "可用电池容量 (kWh)"),
        ("annual_shifted_kwh_est", "全年可转移电量 估计 (kWh/年)"),
        ("max_shiftable_kwh", "可最大转移上限 (kWh/年)"),
        ("final_annual_shifted_kwh", "实际可转移电量 (kWh/年)"),
        ("annual_savings", "年度节省 (AUD/年)"),
        ("total_cost", "项目总成本 (AUD)"),
        ("price_base", "报价基准 (AUD)"),
        ("payback_years", "回本年限 (年)"),
        ("new_self_consumption_rate", "加电池后自用率"),
        ("roi_warning", "回报警告"),
    ]
    r = 2
    for key, cn in retro_params:
        ws_retro[f"A{r}"] = f"{cn}\n({key})"
        r += 1
    sizes = [refs['retrofit_plan_a_kwh'], refs['retrofit_plan_b_kwh'], refs['retrofit_plan_c_kwh'], refs['retrofit_plan_d_kwh']]
    esg = refs['existing_solar_annual_gen_kwh']
    for col_idx in range(2,6):
        col = get_column_letter(col_idx)
        size = sizes[col_idx-2]
        ws_retro[f"{col}2"] = f"={size}"
        ws_retro[f"{col}3"] = f"={col}2*{refs['battery_dod']}"
        ws_retro[f"{col}4"] = f"={col}3*{refs['battery_effective_usage_factor']}*365"
        ws_retro[f"{col}5"] = f"=IF({esg}=\"\", ({refs['roof_max_panels']}*{refs['panel_power_kw']}*0.7*1200)*(1-{refs['existing_sc_rate']}), {esg}*(1-{refs['existing_sc_rate']}))"
        ws_retro[f"{col}6"] = f"=MIN({col}4, {col}5)"
        ws_retro[f"{col}7"] = f"={col}6*({refs['grid_buy_rate']}-{refs['grid_sell_rate']})"
        ws_retro[f"{col}8"] = f"={col}2*{refs['battery_unit_cost_per_kwh']} + {refs['battery_install_base_cost']} + ({col}2*{refs['battery_install_cost_per_kwh']})"
        ws_retro[f"{col}9"] = f"={col}8*(1+{refs['profit_margin_rate']})"
        ws_retro[f"{col}10"] = f"=IF({col}7>0, {col}9/{col}7, \"Inf\")"
        ws_retro[f"{col}11"] = f"=IF(OR({esg}=\"\", {esg}=0), (({refs['roof_max_panels']}*{refs['panel_power_kw']}*0.7*1200)*{refs['existing_sc_rate']} + {col}6)/({refs['roof_max_panels']}*{refs['panel_power_kw']}*0.7*1200), ({esg}*{refs['existing_sc_rate']} + {col}6)/{esg})"
        ws_retro[f"{col}12"] = f"=IF({col}6<500, \"Low ROI - small export available\", \"OK\")"

    ws_readme = wb.create_sheet("Readme_Chinese")
    ws_readme.append(["说明（中文）"])
    ws_readme.append(["此版本为详细版，保留面板/逆变器/电池分项成本与取整逻辑。"])
    for cell in ws_readme["1:1"]:
        cell.font = Font(bold=True)

    for ws in [ws_inputs, ws_plans, ws_retro]:
        for col in range(1, ws.max_column+1):
            max_len = 0
            col_letter = get_column_letter(col)
            for cell in ws[col_letter]:
                try:
                    l = len(str(cell.value))
                except:
                    l = 0
                if l > max_len:
                    max_len = l
            ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(path)
    return path

file1 = "/mnt/data/solar_estimate_recommended.xlsx"
file2 = "/mnt/data/solar_estimate_simplified_hardware_cost.xlsx"
create_detailed_file(file1)
create_simplified_file(file2)
file1, file2

