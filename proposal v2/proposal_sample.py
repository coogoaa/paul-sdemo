# Create two Excel files (detailed and simplified) that use facet_max_power_kw summation for roof_max_power_kw.
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def create_file(path, simplified=False):
    wb = Workbook()
    # Inputs sheet
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    ws_inputs.append(["参数名称 (EN)", "参数中文说明", "默认值", "类型 (上游/预设/计算)", "备注"])
    inputs = [
        ("panel_power_kw", "单块面板功率 (kW)", 0.41, "预设", "面板规格"),
        ("panel_area_m2", "单块面板面积 (m2)", 1.94, "预设", ""),
        ("dc_ac_ratio", "容配比 (DC/AC)", 1.33, "预设", ""),
        ("yield_per_kw_per_year", "每kW年发电量 (kWh/kW/年)", 1460, "预设/兜底", "pvlib 优先"),
        ("baseline_self_consumption_rate", "基线自用率 (无电池)", 0.30, "预设", ""),
        ("plan_a_capacity_factor", "方案A 容量系数", 0.2, "预设", ""),
        ("plan_b_capacity_factor", "方案B 容量系数", 0.5, "预设", ""),
        ("plan_c_capacity_factor", "方案C 容量系数", 0.8, "预设", ""),
        ("plan_d_capacity_factor", "方案D 容量系数", 1.0, "预设", ""),
        ("plan_a_min_kw", "方案A 最小装机下限 (kW)", 3.5, "预设", ""),
        ("plan_c_max_kw", "方案C 最大上限 (kW)", 13.3, "预设", ""),
        ("battery_dod", "电池放电深度 DoD", 0.90, "预设", ""),
        ("battery_rte", "电池往返效率 RTE", 0.90, "预设", ""),
        ("battery_effective_usage_factor", "电池有效使用系数", 0.90, "预设", "每天平均可用比例"),
        ("panel_unit_cost", "单块面板成本 (AUD/块)", 350, "预设", ""),
        ("inverter_unit_cost_per_kw", "逆变器成本 (AUD/kW)", 250, "预设", ""),
        ("battery_unit_cost_per_kwh", "电池成本 (AUD/kWh)", 800, "预设", ""),
        ("install_base_cost", "安装基础费用 (AUD)", 1000, "预设", ""),
        ("install_cost_per_kw", "安装每kW费用 (AUD/kW)", 1000, "预设", ""),
        ("hardware_cost_per_kw", "硬件成本简化 (AUD/kW)", 1200, "预设", "仅简化版使用"),
        ("profit_margin_rate", "利润率", 0.10, "预设", ""),
        ("grid_buy_rate", "购电价 (AUD/kWh)", 0.33, "预设", ""),
        ("grid_sell_rate", "售电/馈网价 (AUD/kWh)", 0.07, "预设", ""),
        ("annual_home_usage_proxy_med", "Med 用户年用电 (kWh)", 6000, "预设", "无账单时兜底"),
        ("existing_sc_rate", "已有系统基线自用率", 0.30, "预设", ""),
        ("retrofit_plan_a_kwh", "Retrofit A 电池容量 (kWh)", 5, "预设", ""),
        ("retrofit_plan_b_kwh", "Retrofit B 电池容量 (kWh)", 10, "预设", ""),
        ("retrofit_plan_c_kwh", "Retrofit C 电池容量 (kWh)", 13.5, "预设", ""),
        ("retrofit_plan_d_kwh", "Retrofit D 电池容量 (kWh)", 20, "预设", ""),
    ]
    row = 2
    refs = {}
    for key, cn, val, typ, note in inputs:
        ws_inputs.append([key, cn, val, typ, note])
        refs[key] = f"C{row}"
        row += 1
    for c in ws_inputs["1:1"]:
        c.font = Font(bold=True)
    # Facets sheet (upstream data)
    ws_facets = wb.create_sheet("Facets")
    ws_facets.append(["facet_id", "facet_area_m2", "facet_max_panels", "facet_max_power_kw"])
    # put example facet rows; user can replace or paste upstream data
    example_facets = [
        (1, 12.0, 6, "=C2*Inputs!$C$2/Inputs!$C$2"),  # placeholder formula will be overwritten by values below
        (2, 8.5, 4, 4*0.41),
        (3, 20.0, 10, 10*0.41),
        (4, 6.0, 3, 3*0.41),
    ]
    # We'll write numeric values for facet_max_power_kw based on facet_max_panels * panel_power_kw for sample rows
    for i, (fid, area, maxp, power) in enumerate(example_facets, start=2):
        # facet_max_power_kw = facet_max_panels * Inputs!panel_power_kw
        ws_facets.append([fid, area, maxp, f"=C{i}*Inputs!{refs['panel_power_kw']}"])
    for c in ws_facets["1:1"]:
        c.font = Font(bold=True)
    # Summary sheet with roof aggregation and plans
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["参数 (EN)", "参数中文", "值/公式", "说明"])
    ws_sum.append(["roof_max_panels", "屋顶最大面板数 (块)", f"=SUM(Facets!C2:C100)", "上游 facet_max_panels 求和（示例范围至 C100）"])
    ws_sum.append(["roof_max_power_kw", "屋顶最大装机容量 (kW)", f"=SUM(Facets!D2:D100)", "上游 facet_max_power_kw 求和（主来源）"])
    # Plans sheet
    ws_plans = wb.create_sheet("Plans")
    ws_plans.append(["参数（中/EN）", "Plan A", "Plan B", "Plan C", "Plan D"])
    for c in ws_plans["1:1"]:
        c.font = Font(bold=True)
    plan_params = [
        ("prelim_kw", "目标连续装机容量 prelim (kW)"),
        ("panel_count", "面板数量 (块)"),
        ("actual_solar_kw", "实际装机容量 (kW)"),
        ("inverter_kw", "逆变器容量 (kW)"),
        ("annual_generation_kwh", "年发电量 (kWh/年)"),
        ("annual_self_consumption_kwh", "年自用电量 (kWh/年)"),
        ("annual_export_kwh", "年上网电量 (kWh/年)"),
        ("annual_export_revenue", "年上网收益 (AUD/年)"),
        ("annual_savings", "年节省金额 (AUD/年)"),
        ("total_cost", "项目总成本 (AUD)"),
        ("price_base", "报价基准 (AUD)"),
        ("payback_base_years", "回本 基线 (年)"),
    ]
    r = 2
    for key, cn in plan_params:
        ws_plans[f"A{r}"] = f"{cn}\n({key})"
        r += 1
    param_row = {plan_params[i][0]: 2+i for i in range(len(plan_params))}
    # Fill formulas for each plan column
    for col_idx, plan in enumerate(["A","B","C","D"], start=2):
        col = get_column_letter(col_idx)
        # prelim = roof_max_power_kw * capacity_factor (A has min and C has max)
        if plan == "A":
            formula_prelim = f"=MAX(MIN(Summary!$B$3*Inputs!{refs['plan_c_max_kw']}, Summary!$B$3*Inputs!{refs['plan_a_capacity_factor']}), Inputs!{refs['plan_a_min_kw']})"
        elif plan == "B":
            formula_prelim = f"=Summary!$B$3*Inputs!{refs['plan_b_capacity_factor']}"
        elif plan == "C":
            formula_prelim = f"=MIN(Summary!$B$3*Inputs!{refs['plan_c_capacity_factor']}, Inputs!{refs['plan_c_max_kw']})"
        else:
            formula_prelim = f"=Summary!$B$3*Inputs!{refs['plan_d_capacity_factor']}"
        ws_plans[f"{col}{param_row['prelim_kw']}"] = formula_prelim
        # panel_count = INT(prelim / panel_power_kw)
        ws_plans[f"{col}{param_row['panel_count']}"] = f"=INT({col}{param_row['prelim_kw']}/Inputs!{refs['panel_power_kw']})"
        # actual_solar_kw = panel_count * panel_power_kw
        ws_plans[f"{col}{param_row['actual_solar_kw']}"] = f"={col}{param_row['panel_count']}*Inputs!{refs['panel_power_kw']}"
        # inverter = CEILING(actual / dc_ac, 0.1)
        ws_plans[f"{col}{param_row['inverter_kw']}"] = f"=CEILING({col}{param_row['actual_solar_kw']}/Inputs!{refs['dc_ac_ratio']},0.1)"
        # annual_generation = actual_solar_kw * yield_per_kw_per_year (pvlib override optional)
        ws_plans[f"{col}{param_row['annual_generation_kwh']}"] = f"={col}{param_row['actual_solar_kw']}*Inputs!{refs['yield_per_kw_per_year']}"
        # annual_self_consumption = MIN(annual_generation * target_sc_rate, annual_home_usage_proxy_med)
        target = f"Inputs!{refs['plan_a_target_sc_rate']}" if plan=="A" else (f"Inputs!{refs['plan_b_target_sc_rate']}" if plan=="B" else (f"Inputs!{refs['plan_c_target_sc_rate']}" if plan=="C" else f"Inputs!{refs['plan_d_target_sc_rate']}"))
        ws_plans[f"{col}{param_row['annual_self_consumption_kwh']}"] = f"=MIN({col}{param_row['annual_generation_kwh']}*{target}, Inputs!{refs['annual_home_usage_proxy_med']})"
        # export = gen - self
        ws_plans[f"{col}{param_row['annual_export_kwh']}"] = f"={col}{param_row['annual_generation_kwh']} - {col}{param_row['annual_self_consumption_kwh']}"
        # export_revenue = export * grid_sell
        ws_plans[f"{col}{param_row['annual_export_revenue']}"] = f"={col}{param_row['annual_export_kwh']}*Inputs!{refs['grid_sell_rate']}"
        # annual_savings = self*grid_buy + export_revenue
        ws_plans[f"{col}{param_row['annual_savings']}"] = f"={col}{param_row['annual_self_consumption_kwh']}*Inputs!{refs['grid_buy_rate']} + {col}{param_row['annual_export_revenue']}"
        # total_cost
        if simplified:
            ws_plans[f"{col}{param_row['total_cost']}"] = f"={col}{param_row['actual_solar_kw']}*Inputs!{refs['hardware_cost_per_kw']} + Inputs!{refs['install_base_cost']} + ({col}{param_row['actual_solar_kw']}*Inputs!{refs['install_cost_per_kw']})"
        else:
            ws_plans[f"{col}{param_row['total_cost']}"] = f"=({col}{param_row['panel_count']}*Inputs!{refs['panel_unit_cost']}) + ({col}{param_row['inverter_kw']}*Inputs!{refs['inverter_unit_cost_per_kw']}) + (Inputs!{refs['install_base_cost']} + ({col}{param_row['actual_solar_kw']}*Inputs!{refs['install_cost_per_kw']}))"
        # price_base
        ws_plans[f"{col}{param_row['price_base']}"] = f"={col}{param_row['total_cost']}*(1+Inputs!{refs['profit_margin_rate']})"
        # payback base = price_base / annual_savings (guard)
        ws_plans[f"{col}{param_row['payback_base_years']}"] = f"=IF({col}{param_row['annual_savings']}>0, {col}{param_row['price_base']}/{col}{param_row['annual_savings']}, \"Inf\")"

    # Battery retrofit sheet
    ws_retro = wb.create_sheet("Battery_Retrofit")
    ws_retro.append(["参数(EN)", "说明", "Retrofit A", "Retrofit B", "Retrofit C", "Retrofit D"])
    for c in ws_retro["1:1"]:
        c.font = Font(bold=True)
    # headers rows for keys
    keys = ["battery_nominal_kwh", "usable_batt_kwh", "annual_shiftable_kwh", "existing_self_kwh", "existing_export_kwh", "final_shifted_kwh", "annual_savings_increase", "retro_total_cost", "retro_price_base", "retro_payback_years", "new_self_consumption_rate", "roi_warning"]
    r = 2
    for k in keys:
        ws_retro[f"A{r}"] = k
        r += 1
    # populate columns with formulas referencing Inputs and Facets
    sizes = [refs['retrofit_plan_a_kwh'], refs['retrofit_plan_b_kwh'], refs['retrofit_plan_c_kwh'], refs['retrofit_plan_d_kwh']]
    esg = f"Inputs!{refs['existing_solar_annual_gen_kwh']}"
    for col_idx in range(3, 7):
        col = get_column_letter(col_idx)
        size_ref = f"Inputs!{sizes[col_idx-3]}"
        # battery_nominal_kwh
        ws_retro[f"{col}2"] = f"={size_ref}"
        # usable_batt = nominal * DoD
        ws_retro[f"{col}3"] = f"={col}2*Inputs!{refs['battery_dod']}"
        # annual_shiftable = usable * usage_factor * 365
        ws_retro[f"{col}4"] = f"={col}3*Inputs!{refs['battery_effective_usage_factor']}*365"
        # existing_self_kwh = IF existing_gen provided use MIN(existing_gen*existing_sc_rate, proxy_med) else estimate from roof_max_power_kw*installed_fraction*yield
        # We'll estimate existing_gen as roof_max_power_kw * 0.7 * yield_per_kw_per_year if blank
        ws_retro[f"{col}5"] = f"=IF({esg}=\"\", MIN(Summary!$B$3*0.7*Inputs!{refs['yield_per_kw_per_year']}*Inputs!{refs['existing_sc_rate']}, Inputs!{refs['annual_home_usage_proxy_med']}), MIN({esg}*Inputs!{refs['existing_sc_rate']}, Inputs!{refs['annual_home_usage_proxy_med']}))"
        # existing_export = total gen - existing_self (total gen uses esg or estimate)
        ws_retro[f"{col}6"] = f"=IF({esg}=\"\", (Summary!$B$3*0.7*Inputs!{refs['yield_per_kw_per_year']}) - {col}5, {esg} - {col}5)"
        # final_shifted = MIN(annual_shiftable, existing_export)
        ws_retro[f"{col}7"] = f"=MIN({col}4, {col}6)"
        # annual_savings_increase = final_shifted * (grid_buy - grid_sell)
        ws_retro[f"{col}8"] = f"={col}7*(Inputs!{refs['grid_buy_rate']} - Inputs!{refs['grid_sell_rate']})"
        # retro_total_cost
        ws_retro[f"{col}9"] = f"={col}2*Inputs!{refs['battery_unit_cost_per_kwh']} + Inputs!{refs['battery_install_base_cost']} + ({col}2*Inputs!{refs['battery_install_cost_per_kwh']})"
        # retro_price_base
        ws_retro[f"{col}10"] = f"={col}9*(1+Inputs!{refs['profit_margin_rate']})"
        # retro_payback_years
        ws_retro[f"{col}11"] = f"=IF({col}8>0, {col}10/{col}8, \"Inf\")"
        # new self consumption rate
        ws_retro[f"{col}12"] = f"=IF(Inputs!{refs['existing_solar_annual_gen_kwh']}=\"\", ( {col}5 + {col}7 ) / (Summary!$B$3*0.7*Inputs!{refs['yield_per_kw_per_year']}), ( {col}5 + {col}7 ) / Inputs!{refs['existing_solar_annual_gen_kwh']})"
        # roi warning
        ws_retro[f"{col}13"] = f"=IF({col}7<500, \"Low ROI - small export available\", \"OK\")"

    # Explanation sheet: very simple language step-by-step for each computation
    ws_explain = wb.create_sheet("Explain_Steps_小学生版")
    ws_explain.append(["步骤", "输入/参数", "这是为什么（小学生能懂的讲法）", "在表里哪里"])
    rows = [
        ("1", "上游：Facets 表（每个有效坡面）", "上游会告诉我们每个屋顶小块（坡面）能装多少块板，和每块小坡面最多能装多大功率。就像把屋顶分成好几个小格子，每格子有自己的容量。", "Facets!C/D 列（facet_max_panels / facet_max_power_kw）"),
        ("2", "把每个小格子的功率加起来", "把每个小格子的最大功率加起来，得到整个屋顶最多能装多少千瓦。就像把所有小格子的气球都充满，算总共能装多少气球。", "Summary!B3 = SUM(Facets!D2:D100)"),
        ("3", "方案目标容量（连续值）", "我们会用一个百分比（容量系数）来决定做多大：比如半满、八成或全部。用屋顶最大功率乘以这个系数，就得到目标容量。", "Plans 表的 prelim_kw 列"),
        ("4", "把目标容量换成整块面板数", "目标容量可能不是整数块板，我们要把它除以每块板的瓦数，再向下取整，得到能装的整块板数。不能装半块板。", "Plans panel_count = INT(prelim / Inputs!panel_power_kw)"),
        ("5", "把面板数换回容量（真实可装容量）", "把整块的面板数乘以单块面板功率，得到实际装机容量。这个才是可以真的装上去的功率。", "Plans actual_solar_kw = panel_count * Inputs!panel_power_kw"),
        ("6", "逆变器容量", "逆变器是把太阳能直流电变成家里用的交流电的机器，它大小要比太阳能略小或相符，用一个比例来算并向上取整到可以买的档位。", "Plans inverter_kw = CEILING(actual_solar_kw / Inputs!dc_ac_ratio, 0.1)"),
        ("7", "年发电量（pvlib 优先）", "如果有地点和天气数据，用 pvlib 模拟更准；没有就用每千瓦年产多少电的默认值乘以装机容量。", "Plans annual_generation_kwh = actual_solar_kw * Inputs!yield_per_kw_per_year (pvlib 覆盖可选)"),
        ("8", "年自用量（我们家自己用掉的太阳电）", "拿年发电量乘以自用率，但不能比家庭一年用的电还多（不会把多的电都吃掉）。", "annual_self = MIN(annual_gen * target_sc_rate, Inputs!annual_home_usage_proxy_med)"),
        ("9", "年上网量（多出来卖给电网的）", "发电减去自用，就是卖给电网的电。", "annual_export = annual_gen - annual_self"),
        ("10", "年上网收益（卖电的钱）", "把卖给电网的电乘以上网价，就知道我们卖电能拿多少钱。", "annual_export_revenue = annual_export * Inputs!grid_sell_rate"),
        ("11", "年节省（自己用省的钱 + 卖电的钱）", "自己用的电是本来要买的，现在不用买了，省的钱是自用量乘以购电价，再加上卖电的钱。", "annual_savings = annual_self * Inputs!grid_buy_rate + annual_export_revenue"),
        ("12", "项目总成本和报价", "把所有要买的东西（面板、逆变器、安装）加起来，再乘上利润率，得到能给客户的报价。", "Plans total_cost / price_base 列"),
        ("13", "回本年限", "用报价除以每年能省下的钱，得到需要多少年才能把钱赚回来。", "payback = price_base / annual_savings"),
        ("14", "电池扩容的特殊计算（边际回本）", "电池只是把本来要卖给电网的电，存起来再用掉，所以电池带来的额外每年省钱=它能每天搬多少电×365×(买电价-卖电价)。回本用的是电池的价格除以这个额外省的钱。", "Battery_Retrofit 页：annual_savings_increase = final_shifted*(grid_buy-grid_sell); payback = price_base / annual_savings_increase"),
    ]
    for r in rows:
        ws_explain.append(r)
    for c in ws_explain["1:1"]:
        c.font = Font(bold=True)
    # small extra notes sheet mapping types
    ws_map = wb.create_sheet("Param_Map")
    ws_map.append(["参数", "类型", "说明（简短）"])
    for key, cn, val, typ, note in inputs:
        ws_map.append([key, typ, cn + ("; " + note if note else "")])
    for c in ws_map["1:1"]:
        c.font = Font(bold=True)

    # autosize columns for main sheets
    for sheet in [ws_inputs, ws_facets, ws_sum, ws_plans, ws_retro, ws_explain, ws_map]:
        for col in range(1, sheet.max_column+1):
            max_len = 0
            col_letter = get_column_letter(col)
            for cell in sheet[col_letter]:
                try:
                    l = len(str(cell.value))
                except:
                    l = 0
                if l > max_len:
                    max_len = l
            sheet.column_dimensions[col_letter].width = max_len + 2

    wb.save(path)
    return path

file1 = "/mnt/data/solar_with_facets_recommended.xlsx"
file2 = "/mnt/data/solar_with_facets_simplified.xlsx"
create_file(file1, simplified=False)
create_file(file2, simplified=True)
file1, file2

